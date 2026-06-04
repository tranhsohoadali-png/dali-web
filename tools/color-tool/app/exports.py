"""
Các tiện ích xuất file cho chương trình gộp:
- build_xlsx: xuất bảng màu ra Excel (.xlsx), mỗi dòng có ô tô màu nền theo HEX.
- build_legend_image: dựng ảnh bản đồ màu kèm bảng chú giải số thứ tự -> mã DALI.

Phần Excel dùng openpyxl (đã copy vào python_embed). Phần ảnh dùng PIL/numpy
(có sẵn). Không phụ thuộc pandas.
"""
import os

import numpy as np
from PIL import Image, ImageDraw, ImageFont

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


def _hex_clean(hex_value):
    return str(hex_value).lstrip('#').upper()


def _hex_to_rgb(hex_value):
    h = _hex_clean(hex_value)
    try:
        return int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    except (ValueError, IndexError):
        return 0, 0, 0


def build_xlsx(colors, out_path):
    """
    colors: list các dòng [stt, "#HEX", dali, percent]
    Tạo file Excel với cột: STT | Màu | HEX | R | G | B | Mã DALI | % diện tích
    Ô cột "Màu" được tô nền đúng màu HEX.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Bang mau DALI"

    headers = ['STT', 'Màu', 'HEX', 'R', 'G', 'B', 'Mã DALI', '% diện tích']
    widths = [6, 10, 12, 6, 6, 6, 14, 12]
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, (title, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=col, value=title)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='2E7D32')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border
        ws.column_dimensions[c.column_letter].width = w

    for i, row in enumerate(colors, start=2):
        stt = row[0]
        hex_value = row[1] if len(row) > 1 else ''
        dali = row[2] if len(row) > 2 else ''
        percent = row[3] if len(row) > 3 else ''
        r, g, b = _hex_to_rgb(hex_value)
        clean = _hex_clean(hex_value)

        values = [stt, '', clean, r, g, b, dali, percent]
        for col, val in enumerate(values, start=1):
            c = ws.cell(row=i, column=col, value=val)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = border
        # tô màu ô "Màu"
        ws.cell(row=i, column=2).fill = PatternFill('solid', fgColor=clean.zfill(6))
        ws.row_dimensions[i].height = 22

    wb.save(out_path)
    return out_path


def _load_font(size):
    for path in (r'C:\Windows\Fonts\arial.ttf', r'C:\Windows\Fonts\segoeui.ttf'):
        try:
            return ImageFont.truetype(path, size)
        except OSError:
            continue
    return ImageFont.load_default()


def build_legend_image(left_image_path, colors, out_path, title='', max_per_col=12):
    """
    Dựng ảnh tải về theo bố cục mẫu ("K496" / "H092 TEST"):
      - Góc trên-trái: ảnh, phía dưới là tên mã (title) chữ lớn.
      - Bên phải: bảng chú giải, tự ĐỘNG CHIA NHIỀU CỘT khi nhiều màu
        (mỗi cột tối đa max_per_col màu). Mỗi dòng gồm
        [số thứ tự] · [ô màu] · [mã DALI].
    colors: [stt, "#HEX", dali, percent]
    """
    n = max(1, len(colors))

    # ---- Kích thước cố định để chú giải luôn rõ, không phụ thuộc số màu ----
    pad = 50
    gap = 40
    row_h = 96
    swatch_w = 250
    swatch_h = 62
    num_col_w = 110
    dali_w = 300
    col_w = num_col_w + swatch_w + 24 + dali_w

    num_font = _load_font(46)
    dali_font = _load_font(46)
    title_font = _load_font(66)

    ncols = (n + max_per_col - 1) // max_per_col
    height_rows = min(n, max_per_col)
    legend_h = height_rows * row_h

    # ---- Ảnh bên trái ----
    base = Image.open(left_image_path).convert('RGB')
    bw, bh = base.size
    img_box_w = 820
    scale = img_box_w / bw
    if bh * scale > legend_h:          # không cao quá khối chú giải
        scale = legend_h / bh
    bw, bh = int(bw * scale), int(bh * scale)
    base = base.resize((bw, bh))
    title_h = 110

    canvas_w = pad + bw + gap + ncols * col_w + (ncols - 1) * gap + pad
    canvas_h = pad + max(legend_h, bh + title_h) + pad
    canvas = Image.new('RGB', (canvas_w, canvas_h), 'white')
    canvas.paste(base, (pad, pad))

    draw = ImageDraw.Draw(canvas)
    if title:
        draw.text((pad + bw / 2, pad + bh + title_h * 0.5), str(title),
                  fill='black', font=title_font, anchor='mm')

    legend_x0 = pad + bw + gap
    for i, row in enumerate(colors):
        stt = row[0]
        hex_value = row[1] if len(row) > 1 else '#000000'
        dali = row[2] if len(row) > 2 else ''
        col = i // max_per_col
        r = i % max_per_col
        cx = legend_x0 + col * (col_w + gap)
        cy = pad + row_h * (r + 0.5)

        # số thứ tự
        draw.text((cx + num_col_w * 0.5, cy), str(stt), fill='black', font=num_font, anchor='mm')
        # ô màu
        sx0 = cx + num_col_w
        sy0 = cy - swatch_h / 2
        draw.rectangle([sx0, sy0, sx0 + swatch_w, sy0 + swatch_h],
                       fill=_hex_to_rgb(hex_value), outline=(180, 180, 180))
        # mã DALI
        draw.text((sx0 + swatch_w + 24, cy), str(dali), fill='black', font=dali_font, anchor='lm')

    canvas.save(out_path)
    return out_path
